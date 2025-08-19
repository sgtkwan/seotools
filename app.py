import os
import json
import tempfile
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify
import pandas as pd
from werkzeug.utils import secure_filename
import google.generativeai as genai
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Config
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'uploads')
OUTPUT_FOLDER = os.getenv('OUTPUT_FOLDER', 'outputs')
ALLOWED_EXTENSIONS = {'csv', 'xls', 'xlsx'}
DEFAULT_BATCH_SIZE = int(os.getenv('DEFAULT_BATCH_SIZE', 200))

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', 16 * 1024 * 1024))  # 16MB max file size
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-here-change-this-in-production')

# Configure Gemini API
google_api_key = os.getenv('GOOGLE_API_KEY')
if not google_api_key:
    raise ValueError("GOOGLE_API_KEY environment variable is required")
genai.configure(api_key=google_api_key)

# Ensure upload and output directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

class KeywordClassificationAgent:
    def __init__(self):
        self.model = genai.GenerativeModel('gemini-2.5-flash')
        self.system_prompt = """You are a keyword classification expert. Your task is to categorize keywords using a JSON spec that supports two column modes: (A) predefined tags and (B) instruction-only columns where you must propose tags.

JSON schema:
* `keywords`: array of strings
* `brands`: array of strings
* `columns`: array where each item is one of:
  - Predefined-tags column: `{ "name": string, "tags": string[] }`
  - Instruction-only column: `{ "name": string, "instructions": string }`

Rules:
1) One tag per column: For each keyword, select exactly ONE tag from each column, or leave blank if none apply.
2) Matching: Use exact or close semantic matching to choose tags.
3) Column isolation: Never reuse tags across columns. Tags belong only to their column.
4) Conservative blanks: If no tag fits well, leave the cell blank.
5) Instruction-only columns: For columns with `instructions` (and no `tags`), first infer a small, coherent set of tags (concise phrases, non-overlapping), guided by the instructions and the provided keywords/brands. Then classify the keywords using ONLY those inferred tags. Do NOT invent wildly granular tags; prefer 3–10 clear options.
6) Do not output any explanations, the inferred tag list, or any extra text.

Output format:
Return ONLY a markdown table with:
* Column 1: "Original keyword"
* Subsequent columns: one per column using each column's `name` as the header
* Cells: the single selected tag for that column, or blank

Be precise, consistent, and avoid free-form text that isn't a tag."""
        
        self.json_data = None
        self.keywords = []
        self.columns = []
    
    def convert_excel_to_json(self, file_path, minified=True):
        """Convert CSV or XLSX file to JSON format for keyword classification."""
        try:
            file_extension = Path(file_path).suffix.lower()
            
            if file_extension == '.csv':
                df = pd.read_csv(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                df = pd.read_excel(file_path)
            else:
                raise ValueError("Unsupported file format. Please use CSV or XLSX files.")
            
            df = df.dropna(how='all')
            columns = df.columns.tolist()
            
            if len(columns) < 2:
                raise ValueError("File must have at least 2 columns (keyword and brand)")
            
            keyword_col = columns[0]
            keywords = df[keyword_col].dropna().astype(str).tolist()
            keywords = [kw.strip() for kw in keywords if kw.strip()]
            
            brand_col = columns[1]
            brands = df[brand_col].dropna().astype(str).tolist()
            brands = [brand.strip() for brand in brands if brand.strip()]
            
            result = {
                "keywords": keywords,
                "brands": list(set(brands)),
                "columns": []
            }
            
            for col in columns[2:]:
                col_values = df[col].dropna().astype(str).tolist()
                col_values = [val.strip() for val in col_values if val.strip()]
                unique_values = list(set(col_values))
                
                if unique_values:
                    column_data = {
                        "name": col,
                        "tags": unique_values
                    }
                else:
                    # No explicit tags provided in the sheet; mark as instruction-only column.
                    column_data = {
                        "name": col,
                        "instructions": ""
                    }
                result["columns"].append(column_data)
            
            return result
            
        except Exception as e:
            raise Exception(f"Error processing file: {str(e)}")
    
    def load_data(self, file_path=None, json_data=None):
        """Load data either from Excel file or directly from JSON."""
        if file_path:
            self.json_data = self.convert_excel_to_json(file_path)
        elif json_data:
            self.json_data = json_data
        else:
            raise ValueError("Either file_path or json_data must be provided")
        
        self.keywords = self.json_data.get("keywords", [])
        self.columns = self.json_data.get("columns", [])
        
        return {
            "keywords_count": len(self.keywords),
            "columns": [col['name'] for col in self.columns],
            "brands": self.json_data.get('brands', [])
        }
    
    def classify_keywords_to_excel(self, output_file, batch_size=DEFAULT_BATCH_SIZE, progress_callback=None):
        """Classify all keywords and save directly to Excel file."""
        if not self.json_data:
            raise ValueError("No data loaded. Use load_data() first.")
        
        total_keywords = len(self.keywords)
        all_data = []
        headers = ['Original keyword'] + [col['name'] for col in self.columns]
        
        for i in range(0, total_keywords, batch_size):
            batch_keywords = self.keywords[i:i + batch_size]
            batch_num = (i // batch_size) + 1
            total_batches = (total_keywords + batch_size - 1) // batch_size
            
            if progress_callback:
                progress_callback(batch_num, total_batches, len(batch_keywords))
            
            batch_json = self.json_data.copy()
            batch_json["keywords"] = batch_keywords
            batch_json_str = json.dumps(batch_json, indent=2, ensure_ascii=False)
            
            prompt = f"""{self.system_prompt}

**JSON Data:**
```json
{batch_json_str}
```

**Task:** Classify these {len(batch_keywords)} keywords according to the rules above. Return ONLY the markdown table with proper headers."""
            
            try:
                response = self.model.generate_content(prompt)
                batch_result = response.text.strip()
                batch_data = self._parse_markdown_to_data(
                    batch_result,
                    expected_keywords=batch_keywords,
                    expected_num_columns=len(headers)
                )
                all_data.extend(batch_data)
                
            except Exception as e:
                # Create fallback data for this batch
                for keyword in batch_keywords:
                    row_data = [keyword] + [''] * len(self.columns)
                    all_data.append(row_data)
        
        self._save_data_to_excel(headers, all_data, output_file)
        return len(all_data)
    
    def _parse_markdown_to_data(self, markdown_result, expected_keywords, expected_num_columns):
        """Parse markdown table to list of lists and ignore stray/non-data rows.

        Rules:
        - Row must have the same number of columns as the header.
        - First cell must match one of the expected keywords (case-insensitive).
        - Ignore header separator and decoration rows.
        """
        normalized_expected = {str(k).strip().lower() for k in expected_keywords}

        lines = [line.strip() for line in markdown_result.split('\n') if '|' in line]

        filtered_lines = []
        for line in lines:
            # Skip typical markdown separator rows or lines that are only pipes/hyphens/colons/spaces
            if not (line.count('-') > len(line) * 0.5 or line.startswith('|--') or
                    all(c in '|-: ' for c in line.replace('|', ''))):
                filtered_lines.append(line)

        if not filtered_lines:
            return []

        # Determine expected column count using header if possible
        header_cells = [c.strip() for c in filtered_lines[0].split('|')[1:-1]]
        header_count = len(header_cells) if header_cells else expected_num_columns

        data_rows = []
        for line in filtered_lines[1:]:  # Skip header
            cells = [cell.strip() for cell in line.split('|')[1:-1]]

            # Column count must match header
            if len(cells) != header_count:
                continue

            # First column must be an expected keyword
            if not cells:
                continue
            first_cell_normalized = cells[0].strip().lower()
            if first_cell_normalized not in normalized_expected:
                continue

            data_rows.append(cells)

        return data_rows
    
    def _save_data_to_excel(self, headers, data_rows, output_file):
        """Save data directly to Excel with basic formatting."""
        try:
            if not output_file.endswith('.xlsx'):
                output_file = f"{output_file.rsplit('.', 1)[0]}.xlsx"
            
            df = pd.DataFrame(data_rows, columns=headers)
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Classification Results', index=False)
                self._format_excel_worksheet(writer, 'Classification Results', headers, len(data_rows))
                
        except Exception as e:
            raise Exception(f"Error creating Excel file: {e}")
    
    def _format_excel_worksheet(self, writer, sheet_name, headers, data_row_count):
        """Apply formatting to Excel worksheet."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            worksheet = writer.sheets[sheet_name]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, column_title in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        cell_length = len(str(cell.value)) if cell.value else 0
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=data_row_count+1, 
                                          min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        
        except Exception as e:
            print(f"Error formatting Excel worksheet: {e}")

def allowed_file(filename):
    """Check if file has allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/", methods=["GET", "POST"])
def home():
    """Main route for file upload and preview."""
    if request.method == "GET":
        # Provide the current/default system prompt to prefill the textarea on the index page
        default_prompt_agent = KeywordClassificationAgent()
        return render_template("index.html", default_system_prompt=default_prompt_agent.system_prompt)
    
    file = request.files.get("file")
    
    if not file or file.filename == '':
        flash("No file selected", "error")
        return redirect(url_for('home'))
    
    if not allowed_file(file.filename):
        flash("Unsupported file type. Please upload CSV, XLS, or XLSX files.", "error")
        return redirect(url_for('home'))
    
    try:
        # Save uploaded file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        # Initialize classification agent and load data
        agent = KeywordClassificationAgent()
        summary = agent.load_data(file_path=filepath)
        
        # Store file info in session-like manner (you might want to use actual sessions)
        file_info = {
            'filepath': filepath,
            'original_filename': filename,
            'summary': summary
        }
        
        # Capture user-provided system prompt (fallback to default if missing)
        user_system_prompt = request.form.get('system_prompt') or agent.system_prompt
        
        return render_template("preview.html", 
                             file_info=file_info,
                             unique_filename=unique_filename,
                             system_prompt=user_system_prompt,
                             columns_detail=agent.columns)
        
    except Exception as e:
        flash(f"Error processing file: {str(e)}", "error")
        return redirect(url_for('home'))

@app.route("/process/<filename>", methods=["POST"])
def process_file(filename):
    """Process the uploaded file with Gemini and return Excel result."""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(filepath):
            flash("File not found", "error")
            return redirect(url_for('home'))
        
        # Get batch size from form (with default)
        batch_size = int(request.form.get('batch_size', DEFAULT_BATCH_SIZE))
        # Get the (possibly edited) system prompt from the form
        user_system_prompt = request.form.get('system_prompt')
        
        # Initialize agent and load data
        agent = KeywordClassificationAgent()
        if user_system_prompt:
            agent.system_prompt = user_system_prompt
        agent.load_data(file_path=filepath)

        # Collect any instruction inputs for instruction-only columns and apply to agent.columns
        updated_columns = []
        for index, col in enumerate(agent.columns):
            # Expect hidden field instruction_col_name_{i} for instruction columns as rendered in preview
            instruction_name_field = f"instruction_col_name_{index}"
            instruction_value_field = f"instruction_{index}"
            provided_col_name = request.form.get(instruction_name_field)
            provided_instruction = request.form.get(instruction_value_field, "").strip()

            if isinstance(col, dict) and 'tags' not in col:
                # Instruction-only column. If provided, set instructions text
                column_copy = dict(col)
                if provided_col_name == col.get('name') and provided_instruction:
                    column_copy['instructions'] = provided_instruction
                updated_columns.append(column_copy)
            else:
                updated_columns.append(col)

        agent.columns = updated_columns
        
        # Create output filename
        base_name = Path(filename).stem
        output_filename = f"classified_{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_filepath = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
        
        # Process with Gemini
        def progress_callback(batch_num, total_batches, batch_size):
            # In a real application, you might want to use WebSockets for real-time progress
            print(f"Processing batch {batch_num}/{total_batches} ({batch_size} keywords)")
        
        processed_count = agent.classify_keywords_to_excel(
            output_filepath, 
            batch_size=batch_size,
            progress_callback=progress_callback
        )
        
        # Clean up uploaded file
        os.remove(filepath)
        
        # Return success page with download link instead of direct download
        return render_template('success.html', 
                             output_filename=output_filename,
                             processed_count=processed_count,
                             original_filename=Path(filename).stem)
        
    except Exception as e:
        flash(f"Error processing file: {str(e)}", "error")
        return redirect(url_for('home'))

@app.route("/download/<filename>")
def download_file(filename):
    """Secure download endpoint with proper headers."""
    try:
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            flash("File not found or expired", "error")
            return redirect(url_for('home'))
        
        # Create response with secure headers
        response = send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Add security headers to prevent Chrome warnings
        response.headers['Content-Security-Policy'] = "default-src 'self'"
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        
        # Clean up file after 5 seconds (optional)
        def cleanup_file():
            import threading
            import time
            def delayed_cleanup():
                time.sleep(5)
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except:
                    pass
            thread = threading.Thread(target=delayed_cleanup)
            thread.daemon = True
            thread.start()
        
        cleanup_file()
        return response
        
    except Exception as e:
        flash(f"Error downloading file: {str(e)}", "error")
        return redirect(url_for('home'))

@app.route("/progress/<filename>")
def get_progress(filename):
    """Get processing progress (placeholder for future WebSocket implementation)."""
    # This is a placeholder for real-time progress tracking
    # In a production app, you'd implement this with WebSockets or Server-Sent Events
    return jsonify({"status": "processing", "progress": 50})

@app.errorhandler(413)
def too_large(e):
    flash("File is too large. Maximum size is 16MB.", "error")
    return redirect(url_for('home'))

if __name__ == "__main__":
    # Get port from environment variable (Render sets PORT)
    port = int(os.getenv('PORT', 5000))
    
    # Only run in debug mode if explicitly set
    debug = os.getenv('FLASK_ENV') == 'development'
    
    app.run(debug=debug, host='0.0.0.0', port=port)